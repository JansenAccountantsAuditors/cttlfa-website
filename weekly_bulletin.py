# -*- coding: utf-8 -*-
import json, datetime, base64, html, re, os, shutil, urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SITE='https://www.cttfa.co.za'
def _get(url, timeout=30):
    req=urllib.request.Request(url, headers={'User-Agent':'CTTLFA-weekly-bulletin'})
    return urllib.request.urlopen(req, timeout=timeout).read().decode('utf-8','ignore')

# --- season data: live from the site, with a local fallback so the job never dies ---
try:
    J=json.load(open(os.environ.get('SEASON_JSON','season.json'))); SEASON_SRC='local season.json'
except Exception:
    J=json.loads(_get(SITE+'/season.json')); SEASON_SRC='live site'
LG=J['leagues']; YR=int(J.get('season') or 2026)
CRESTS=J.get('crests') or {}
CRESTBASE=SITE+'/'+(J.get('crestBase') or 'photos/crests')

MON={'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
def d2(s):
    try:
        p=str(s).split(); return datetime.date(YR,MON[p[1][:3].lower()],int(p[0]))
    except: return None
def clean_venue(v):
    v=(v or '').strip()
    if v.lower() in ('none','bye'): return ''
    prev=None
    while prev!=v:
        prev=v; v=re.sub(r'^(.+?)\s+\1(\s|$)', r'\1\2', v)
    return v.strip()

# --- mode: 'weekend'   = Thursday bulletin (this weekend Fri-Sun, with referees)
#          'weekahead' = Monday bulletin (whole current week Mon-Sun, no referees yet) ---
MODE=(os.environ.get('BULLETIN_MODE','weekend') or 'weekend').strip().lower()
if MODE not in ('weekend','weekahead'): MODE='weekend'
SHOW_REFS=(MODE=='weekend')

def _run_date():
    ov=os.environ.get('RUN_DATE','').strip()
    if ov:
        y,m,d=map(int,ov.split('-')); return datetime.date(y,m,d)
    return datetime.date.today()
RUN=_run_date()
if MODE=='weekahead':
    # the whole current week, Monday to Sunday (the job runs Monday 12:00 SAST)
    _MON=RUN-datetime.timedelta(RUN.weekday())
    _WD=[_MON+datetime.timedelta(i) for i in range(7)]
    _FRI,_SAT,_SUN=_WD[4],_WD[5],_WD[6]
    _RA,_RB=_WD[0],_WD[6]
else:
    # coming weekend = the Fri/Sat/Sun around the next Saturday on/after the run date
    _SAT=RUN+datetime.timedelta((5-RUN.weekday())%7)
    _FRI=_SAT-datetime.timedelta(1); _SUN=_SAT+datetime.timedelta(1)
    _WD=[_FRI,_SAT,_SUN]
    _RA,_RB=_FRI,_SUN
def _lab(d): return '%s %d %s'%(d.strftime('%A'), d.day, d.strftime('%B'))
WEEK={d:_lab(d) for d in _WD}
DAY_ORDER=[_lab(d) for d in _WD]
DAYSHORT={_lab(d):'%s %d'%(d.strftime('%a'), d.day) for d in _WD}
if _RA.month==_RB.month:
    RANGE_LABEL='%d – %d %s %d'%(_RA.day,_RB.day,_RB.strftime('%B'),_RB.year)
else:
    RANGE_LABEL='%d %s – %d %s %d'%(_RA.day,_RA.strftime('%B'),_RB.day,_RB.strftime('%B'),_RB.year)
ISSUE_LABEL='%d %s %d'%(RUN.day,RUN.strftime('%b'),RUN.year)
def is_senior(n): return ('Under' not in n) and ('Girls' not in n)

# ---------- unified weekend fixtures (leagues + cups, deduped) ----------
uni={}
def add(key,row):
    if key in uni:
        if row['type']=='Cup' and uni[key]['type']!='Cup': uni[key]=row
    else: uni[key]=row
for L in LG.values():
    nm=L.get('name','')
    for f in L.get('fixtures',[]):
        dt=d2(f[2] or '')
        if dt in WEEK and f[0] and f[1]:
            add((f[0],f[1],dt,f[3] or ''),{'dt':dt,'day':WEEK[dt],'time':f[3] or 'TBC','comp':nm,'type':'League',
                 'home':f[0],'away':f[1],'venue':clean_venue(f[4]),'senior':is_senior(nm)})
for c in J.get('cups',[]):
    cn=c.get('name','Cup'); grp=c.get('group','')
    senior=not str(grp).startswith('Under')
    for col in c.get('bracket',{}).get('cols',[]):
        for b in col:
            a,bb=b.get('a'),b.get('b')
            def sc(x): return x and x.get('s') not in ('',None)
            if b.get('d') and d2(b['d']) in WEEK and a and a.get('n') and bb and bb.get('n') and not(sc(a) and sc(bb)):
                dt=d2(b['d'])
                add((a['n'],bb['n'],dt,b.get('t') or ''),{'dt':dt,'day':WEEK[dt],'time':b.get('t') or 'TBC','comp':cn,'type':'Cup',
                     'home':a['n'],'away':bb['n'],'venue':clean_venue(b.get('v')),'senior':senior})
rows=list(uni.values())
total=len(rows); sen=sum(1 for r in rows if r['senior']); jun=total-sen
NCOMP=len({r['comp'] for r in rows}); ncup=sum(1 for r in rows if r['type']=='Cup')

# ---------- featured (senior marquee competitions) ----------
# Headline cards shown in the email body = Premier Division + ALL cup/knockout ties only.
# First/Second Division and Women's Premier LEAGUE games are intentionally NOT carded here
# (they stay in the attached spreadsheet + on the website) to keep the email short.
FEATURE={'Premier Division':('Premier Division',False),
         'Premier League Cup':('Premier League Cup',True),
         'First Division Cup':('First Division Cup',True),
         'Second Division Cup':('Second Division Cup',True),
         'Womens Premier League Cup':("Women's Premier Cup",True)}
feat={d:[] for d in WEEK.values()}  # populated after appointments load (needs find_appt)

# ---------- results / standings / logs ----------
def league(name):
    for L in LG.values():
        if L.get('name')==name: return L
pd=league('Premier Division')
results=[(r[0],r[2],r[3],r[1]) for r in (pd.get('results') or [])[:6]]
ptable=[(r[0],r[5]-r[6],r[7]) for r in (pd.get('table') or [])[:4]]
def leadr(n):
    L=league(n); return L['table'][0][0] if (L and L.get('table')) else '—'
first_lead=leadr('First Division'); second_lead=leadr('Second Division')
AGES=[12,14,16,18]
def findlog(label,n):
    for L in LG.values():
        if (L.get('name') or '')=='Under %d %s'%(n,label): return L
def build_tier(label):
    agg={}
    for n in AGES:
        L=findlog(label,n)
        if not L or not L.get('table'): continue
        for r in L['table']:
            a=agg.setdefault(r[0],[0,0]); a[0]+=r[5]-r[6]; a[1]+=r[7]
    return sorted(agg.items(), key=lambda kv:(-kv[1][1],-kv[1][0]))
TIERS=[('Premier One','Premier One',1,4),('Premier Two','Premier Two',2,4),
       ('Premier Three A','Premier Three',1,2),('Premier Three B','Premier Three B',1,2)]

# --- club canonicalisation: read the website's OWN mapping (single source of truth) ---
_FALLBACK_SUFFIX=r'\s+(?:[A-Z]|\d{1,2}|I{2,3}|Reserves?)\.?$'
_FALLBACK_ALIAS={"CR Vasco Da Gama":"CR Vasco da Gama","C.R. Vasco Da Gama":"CR Vasco da Gama","Bellville City":"Bellville City FC","Bellville City FC":"Bellville City FC","Table View":"Table View FC","Sunningdale City":"Sunningdale City FC","Kuilsriver":"Kuilsrivier AFC","FC Kapstadt":"FC Kapstadt","Mutual":"Mutual","Mutual FC":"Mutual","Lansdowne":"Lansdowne FC","Queens Park":"Queens Park FC","West End United":"West End United","Avendale Athletico":"Avendale Athletico","Chelsea Bridgetown":"Chelsea Bridgetown","Clyde Pinelands":"Clyde Pinelands","YSD Macassar":"YSD Macassar","YMO FC":"YMO St Lukes","Saxon Rovers":"Saxon Rovers","Green Point Salesians":"Green Point Salesians","Garlandale FC":"Garlandale","Shosholoza FC":"Shosholoza","Edgemead G/W":"Edgemead Goodwood","Varsity College":"Emeris"}
def load_club_mapping():
    try:
        h=open('index.html',encoding='utf-8').read() if os.path.exists('index.html') else _get(SITE+'/index.html')
        ms=re.search(r'const TEAMSUFFIX=/(.+?)/;',h)
        ma=re.search(r'const CLUBALIAS=(\{.*?\});',h)
        if not ma: raise ValueError('CLUBALIAS not found')
        return re.compile(ms.group(1) if ms else _FALLBACK_SUFFIX), json.loads(ma.group(1)), 'live site'
    except Exception as e:
        return re.compile(_FALLBACK_SUFFIX), dict(_FALLBACK_ALIAS), 'baked fallback (%s)'%e
TEAMSUFFIX, CLUBALIAS, MAP_SRC = load_club_mapping()
def foldclub(t0):
    n=re.sub(r'\.+$','',str(t0 or '').strip()); prev=None
    while prev!=n:
        prev=n; n=re.sub(r'\.+$','',TEAMSUFFIX.sub('',n)).strip()
    return n
def canonclub(t0):
    b=foldclub(t0); return CLUBALIAS.get(b,b)
def crest_url(team):
    for k in (team, foldclub(team), canonclub(team)):
        if k in CRESTS: return CRESTBASE+'/'+str(CRESTS[k])+'/115.jpg'
    return None
NCLUBS=len({canonclub(r['home']) for r in rows}|{canonclub(r['away']) for r in rows})

# --- referee appointments: read from the site's own published appointments.json ---
#     (refreshed on the site every ~2h by the Referee appointments sync; the mailer
#      reads this committed file and no longer pulls the dashboard/Supabase itself).
def ap_date(s):
    s=str(s or '').strip()
    m=re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)   # ISO from the Supabase pull, e.g. 2026-08-15
    if m:
        try: return datetime.date(int(m.group(1)),int(m.group(2)),int(m.group(3)))
        except: return None
    p=s.split()                                      # legacy "Sat 15 Aug" from the browser scrape
    for i,tok in enumerate(p):
        if tok.isdigit():
            try: return datetime.date(YR,MON[p[i+1][:3].lower()],int(tok))
            except: return None
    return None
try:
    _AP=json.load(open('appointments.json'))
    AP_ROWS=(_AP.get('senior') or [])+(_AP.get('junior') or [])+(_AP.get('ladies') or [])
    AP_SRC='site appointments.json: %s'%(_AP.get('week_label') or 'cttfa.co.za')
except Exception as e:
    AP_ROWS=[]; AP_SRC='none (%s)'%e
APPTS=[]
for _r in AP_ROWS:
    try:
        div,dt_s,tm,h,a,ref,a1,a2=_r
        APPTS.append({'dt':ap_date(dt_s),'time':(tm or '').strip(),'home':h,'away':a,'ref':ref,'a1':a1,'a2':a2})
    except Exception: pass
def _norm(s):
    s=re.sub(r'\(guest\)','',str(s or ''),flags=re.I)
    return re.sub(r'[^a-z0-9]','',s.lower())
def _tm(a,b):
    na,nb=_norm(a),_norm(b)
    if not na or not nb: return False
    if na==nb or na.startswith(nb) or nb.startswith(na) or na in nb or nb in na: return True
    ca=_norm(canonclub(re.sub(r'\(guest\)','',str(a),flags=re.I).strip()))
    cb=_norm(canonclub(re.sub(r'\(guest\)','',str(b),flags=re.I).strip()))
    return bool(ca and cb and ca==cb)
def find_appt(r):
    # 1) same orientation (home->home, away->away); 2) reversed pairing (sources
    #    sometimes disagree on which side is home — the referee trio is the same).
    cand=[p for p in APPTS if p['dt']==r['dt'] and _tm(p['home'],r['home']) and _tm(p['away'],r['away'])]
    if not cand:
        cand=[p for p in APPTS if p['dt']==r['dt'] and _tm(p['home'],r['away']) and _tm(p['away'],r['home'])]
    if not cand: return ('','','')
    exact=[p for p in cand if p['time']==(r['time'] or '').strip()]
    p=(exact or cand)[0]
    return (p['ref'] or '', p['a1'] or '', p['a2'] or '')
def clean_ref(x):
    x=(x or '').strip()
    return '' if x.upper() in ('TBA','') else x
AP_MATCHED=sum(1 for r in rows if clean_ref(find_appt(r)[0]))
STAMP=(datetime.datetime.utcnow()+datetime.timedelta(hours=2)).strftime('%H:%M on %A %d %B %Y')+' (SAST)'

# ---------- populate featured cards now that appointments are loaded ----------
for r in rows:
    if r['comp'] in FEATURE:
        lab,iscup=FEATURE[r['comp']]
        ref,a1,a2=find_appt(r)
        feat[r['day']].append((r['time'],r['home'],r['away'],r['venue'],lab,iscup,ref,a1,a2))
for d in feat: feat[d].sort(key=lambda x:x[0])

# ================= XLSX =================
def make_xlsx(path):
    NAVY='FF071A4A'; thin=Side(style='thin',color='FFE6EAF3')
    order={lab:i for i,lab in enumerate(DAY_ORDER)}
    def style_header(ws):
        for c in ws[1]:
            c.font=Font(bold=True,color='FFFFFFFF',size=11); c.fill=PatternFill('solid',fgColor=NAVY)
            c.alignment=Alignment(horizontal='left',vertical='center'); c.border=Border(bottom=Side(style='thin',color='FF999999'))
    wb=openpyxl.Workbook()
    def apptxt(x): return '' if (x or '').strip().upper() in ('TBA','') else x.strip()
    refcols=['Referee','Asst 1','Asst 2'] if SHOW_REFS else []
    ws=wb.active; ws.title='By Club'
    ws.append(['Club','Team','H/A','Opponent','Date','Day','Kick-off','Type','Competition','Venue']+refcols); style_header(ws)
    lr=[]
    for r in rows:
        lr.append((canonclub(r['home']),r['home'],'Home',r['away'],r))
        lr.append((canonclub(r['away']),r['away'],'Away',r['home'],r))
    lr.sort(key=lambda x:(x[0].lower(),order[x[4]['day']],x[4]['time']))
    for i,(club,team,ha,opp,r) in enumerate(lr):
        base=[club,team,ha,opp,DAYSHORT[r['day']],r['day'].split()[0],r['time'],r['type'],r['comp'],r['venue']]
        if SHOW_REFS:
            rf,ra1,ra2=find_appt(r); base+=[apptxt(rf),apptxt(ra1),apptxt(ra2)]
        ws.append(base)
        rr=ws[ws.max_row]
        for c in rr:
            c.alignment=Alignment(vertical='center'); c.border=Border(bottom=thin)
            if i%2: c.fill=PatternFill('solid',fgColor='FFF7F9FC')
        rr[0].font=Font(bold=True,color='FF071A4A')
        rr[2].font=Font(bold=True,color=('FF1a7f37' if ha=='Home' else 'FF8a5a00'))
        if SHOW_REFS: rr[10].font=Font(bold=True,color='FF071A4A')
        if r['type']=='Cup':
            rr[7].font=Font(bold=True,color='FF8A6D0A'); rr[7].fill=PatternFill('solid',fgColor='FFFBF1D6')
        else: rr[7].font=Font(color='FF123FB5')
    widths=[22,22,7,22,11,9,9,8,24,28]+([18,16,16] if SHOW_REFS else [])
    for i,w in enumerate(widths): ws.column_dimensions[chr(65+i)].width=w
    ws.freeze_panes='A2'; ws.auto_filter.ref='A1:%s%d'%(chr(64+len(widths)),len(lr)+1)
    ws2=wb.create_sheet('All Fixtures')
    ws2.append(['Date','Day','Kick-off','Type','Competition','Home','Away','Venue']+refcols); style_header(ws2)
    srt=sorted(rows,key=lambda r:(order[r['day']],r['time'],r['comp']))
    for i,r in enumerate(srt):
        base=[DAYSHORT[r['day']],r['day'].split()[0],r['time'],r['type'],r['comp'],r['home'],r['away'],r['venue']]
        if SHOW_REFS:
            rf,ra1,ra2=find_appt(r); base+=[apptxt(rf),apptxt(ra1),apptxt(ra2)]
        ws2.append(base)
        rr=ws2[ws2.max_row]
        for c in rr:
            c.alignment=Alignment(vertical='center'); c.border=Border(bottom=thin)
            if i%2: c.fill=PatternFill('solid',fgColor='FFF7F9FC')
        if SHOW_REFS: rr[8].font=Font(bold=True,color='FF071A4A')
        if r['type']=='Cup':
            rr[3].font=Font(bold=True,color='FF8A6D0A'); rr[3].fill=PatternFill('solid',fgColor='FFFBF1D6')
        else: rr[3].font=Font(color='FF123FB5')
    widths2=[11,9,9,8,26,22,22,30]+([18,16,16] if SHOW_REFS else [])
    for i,w in enumerate(widths2): ws2.column_dimensions[chr(65+i)].width=w
    ws2.freeze_panes='A2'; ws2.auto_filter.ref='A1:%s%d'%(chr(64+len(widths2)),len(srt)+1)
    wb.save(path)
    return len(lr),len(srt)

# ================= HTML =================
NAVY='#071A4A'; BLUE='#123FB5'; GOLD='#F4B41A'; INK='#0B1220'; MUT='#5A667C'; LINE='#E6EAF3'
GN='#E7F6EC'; GNT='#1a7f37'; RD='#FDEBEC'; RDT='#c0392b'; SOFT='#F5F7FB'; PAD='48'
def esc(s): return html.escape(str(s))
def initials(name):
    return (''.join(w[0] for w in re.sub(r'[^A-Za-z0-9 ]',' ',str(name)).split()[:2]).upper() or '?')
def badge(name):
    u=crest_url(name)
    if u:
        return f'<img src="{u}" width="26" height="26" alt="{esc(initials(name))}" style="width:26px;height:26px;border-radius:50%;border:0;vertical-align:middle;background:#eef2fb;">'
    return f'<span style="display:inline-block;width:26px;height:26px;border-radius:50%;background:#eef2fb;color:{BLUE};font-size:9px;font-weight:bold;line-height:26px;text-align:center;">{esc(initials(name))}</span>'
def team_row(name, home=False):
    tag=f'<span style="color:{MUT};font-size:10px;text-transform:uppercase;letter-spacing:.5px;font-weight:bold;">Home</span>' if home else ''
    return (f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>'
      f'<td width="34" valign="middle" style="width:34px;">{badge(name)}</td>'
      f'<td valign="middle" style="color:{INK};font-size:15px;font-weight:bold;line-height:1.25;">{esc(name)}</td>'
      f'<td align="right" valign="middle">{tag}</td></tr></table>')
def section_banner(title,url,cta='View on the website'):
    return (f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" bgcolor="{NAVY}" style="background:{NAVY};border-radius:10px;"><tr><td bgcolor="{NAVY}" style="background:{NAVY};padding:14px 20px;">'
      f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>'
      f'<td valign="middle" style="color:#ffffff;font-size:14px;font-weight:bold;text-transform:uppercase;letter-spacing:1.5px;">'
      f'<span style="display:inline-block;width:4px;height:15px;background:{GOLD};border-radius:2px;vertical-align:middle;margin-right:12px;"></span>'
      f'<a href="{url}" target="_blank" style="color:#ffffff;text-decoration:none;">{esc(title)}</a></td>'
      f'<td align="right" valign="middle" style="white-space:nowrap;">'
      f'<a href="{url}" target="_blank" style="color:{GOLD};font-size:12px;font-weight:bold;text-decoration:none;">{esc(cta)} &rarr;</a></td>'
      f'</tr></table></td></tr></table>')
def ref_html(ref,a1,a2):
    ref=clean_ref(ref)
    if not ref: return ''
    assts=[x.strip() for x in (a1,a2) if (x or '').strip() and x.strip().upper()!='TBA']
    line=(f'<span style="color:{NAVY};font-weight:bold;">Referee</span> {esc(ref)}')
    if assts:
        line+=f'<span style="color:{MUT};"> &middot; Asst: {esc(", ".join(assts))}</span>'
    return (f'<div style="margin-top:9px;padding-top:9px;border-top:1px dashed {LINE};font-size:11.5px;color:{INK};">'
            f'<span style="display:inline-block;background:#eef2fb;color:{BLUE};font-size:9px;font-weight:bold;text-transform:uppercase;letter-spacing:.4px;padding:2px 7px;border-radius:4px;vertical-align:middle;margin-right:8px;">Match Officials</span>'
            f'{line}</div>')
def fcard(time,home,away,venue,lab,iscup,ref='',a1='',a2=''):
    tagbg=GOLD if iscup else BLUE; tagfg=INK if iscup else '#ffffff'
    star='<span style="color:#8A6D0A;">&#9733;</span> ' if iscup else ''
    v=esc(venue) if venue else ''
    meta=(f'<span style="font-weight:800;color:{NAVY};font-size:15px;">{esc(time)}</span>'
          f'&nbsp;&nbsp;<span style="display:inline-block;background:{tagbg};color:{tagfg};font-size:10px;font-weight:bold;text-transform:uppercase;letter-spacing:.4px;padding:3px 9px;border-radius:5px;">{star}{esc(lab)}</span>'
          + (f'<br><span style="color:{MUT};font-size:11.5px;">{v}</span>' if v else ''))
    return (f'<a href="{SITE}/thisweekend" target="_blank" style="text-decoration:none;color:inherit;display:block;">'
      f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border:1px solid {LINE};{("border-top:3px solid "+GOLD+";") if iscup else ""}border-radius:14px;margin-bottom:12px;background:#ffffff;">'
      f'<tr><td style="padding:14px 16px;">{team_row(home,True)}'
      f'<div style="height:9px;font-size:0;line-height:0;">&nbsp;</div>{team_row(away,False)}'
      f'<div style="border-top:1px solid {LINE};margin-top:12px;padding-top:11px;">{meta}</div>'
      f'{ref_html(ref,a1,a2) if SHOW_REFS else ""}'
      f'</td></tr></table></a>')
def grid3(cards):
    out='<table role="presentation" width="100%" cellpadding="0" cellspacing="0">'
    for i in range(0,len(cards),3):
        chunk=cards[i:i+3]+['']*(3-len(cards[i:i+3]))
        out+='<tr>'+''.join(f'<td class="col3" width="33.33%" valign="top" style="padding:0 7px;">{c}</td>' for c in chunk)+'</tr>'
    return out+'</table>'
def day_section(day):
    cards=[fcard(*c) for c in feat.get(day,[])]
    if not cards: return ''
    return (f'<tr><td class="px" style="padding:20px {PAD}px 0;"><div style="color:{NAVY};font-size:15px;font-weight:bold;padding-bottom:13px;">{esc(day)}</div>{grid3(cards)}</td></tr>')
def logs_table(disp,label,up,down):
    r=build_tier(label); n=len(r)
    head=(f'<tr style="background:{SOFT};color:{MUT};font-size:10.5px;text-transform:uppercase;letter-spacing:.4px;">'
          '<td style="padding:8px 6px 8px 14px;">#</td><td style="padding:8px 6px;">Club</td>'
          '<td align="center" style="padding:8px 6px;">GD</td><td align="right" style="padding:8px 14px 8px 6px;">Pts</td></tr>')
    body=''
    for i,(club,(gd,pts)) in enumerate(r):
        bg=GN if i<up else (RD if i>=n-down else ('#ffffff' if i%2==0 else '#fafbfe'))
        bold='font-weight:bold;' if i<up else ''
        body+=(f'<tr style="background:{bg};"><td style="padding:7px 6px 7px 14px;{bold}">{i+1}</td><td style="padding:7px 6px;{bold}">{esc(club)}</td>'
               f'<td align="center" style="padding:7px 6px;">{gd:+d}</td><td align="right" style="padding:7px 14px 7px 6px;font-weight:bold;">{pts}</td></tr>')
    return (f'<div style="color:{NAVY};font-size:15px;font-weight:bold;padding-bottom:9px;">{esc(disp)} <span style="color:{MUT};font-size:11px;font-weight:normal;">&middot; {up} up, {down} down</span></div>'
            f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:13px;color:{INK};border:1px solid {LINE};border-radius:10px;overflow:hidden;">{head}{body}</table>')
results_html=''.join(f'<tr><td style="padding:9px 0;border-bottom:1px solid {LINE};font-size:14px;color:{INK};">{esc(h)} <b>{esc(hs)} &ndash; {esc(a_)}</b> {esc(a)}</td></tr>' for(h,hs,a_,a) in results)
ptable_html=''.join(f'<tr><td style="padding:8px 0;border-top:1px solid {LINE};{("font-weight:bold;" if i==0 else "")}">{i+1}</td><td style="padding:8px 0;border-top:1px solid {LINE};{("font-weight:bold;" if i==0 else "")}">{esc(c)}</td><td align="center" style="padding:8px 0;border-top:1px solid {LINE};">{gd:+d}</td><td align="right" style="padding:8px 0;border-top:1px solid {LINE};font-weight:bold;">{p}</td></tr>' for i,(c,gd,p) in enumerate(ptable))

def stat(n,lab,gold=False):
    return (f'<td width="20%" align="center" bgcolor="{NAVY}" style="background:{NAVY};padding:22px 6px;{"" if lab=="Matches" else "border-left:1px solid rgba(255,255,255,.14);"}">'
            f'<div class="stat-n" style="color:{GOLD if gold else "#fff"};font-size:34px;font-weight:bold;line-height:1;">{n}</div>'
            f'<div style="color:#9fb0d8;font-size:10.5px;font-weight:bold;text-transform:uppercase;letter-spacing:.7px;padding-top:6px;">{lab}</div></td>')

def page(crest):
    days_html=''.join(day_section(d) for d in DAY_ORDER)

    # ---- mode-specific copy ----
    if MODE=='weekahead':
        KICKER='The Week Ahead'
        EYEBROW='Week Ahead &middot; %s'%RANGE_LABEL
        H1='The week ahead at a glance'
        INTRO=('Dear Clubs, Life Members and Referees,<br><br>Here is the week ahead across the association, '
               'together with a recap of last weekend&rsquo;s results. Every fixture from Monday to Sunday is listed '
               'below and attached as a spreadsheet. This is a summary &mdash; tap any section heading or fixture to open '
               f'the full, live detail on <a href="{SITE}" target="_blank" style="color:{BLUE};font-weight:bold;text-decoration:none;">the website</a>, '
               'which is always the main source. Referee appointments are confirmed midweek and will be published in '
               'Thursday&rsquo;s weekend bulletin.')
        PREHEADER=('The week ahead across the association &mdash; last weekend&rsquo;s results plus every fixture from Monday to '
                   f'Sunday ({sen} senior, {jun} junior, {ncup} cup ties). Referee appointments follow on Thursday.')
        FIX_TITLE="This week's fixtures"
        FIX_NOTE=('Referee appointments are confirmed midweek and will appear on each fixture in '
                  'Thursday&rsquo;s weekend bulletin.')
        HEADLINE_NOTE=('The email shows the Premier Division and cup ties.')
        FIX_ATTACH=('First Division, Second Division and Women&rsquo;s Premier, along with Reserves, 3rd&ndash;6th Divisions, '
                    'Veterans and all junior fixtures, are on the website and in the <b style="color:'+INK+';">attached '
                    'spreadsheet</b> &mdash; open it in Excel and filter to your club to see all your games, home and away.')
        FIX_CTA='See the full week&rsquo;s fixtures on the website'
        RES_TITLE="Last weekend's results & standings"
        RES_HEAD='Last weekend&rsquo;s results &middot; Premier'
    else:
        KICKER='Weekly Club Bulletin'
        EYEBROW='Match Week &middot; %s'%RANGE_LABEL
        H1='Your weekend at a glance'
        INTRO=('Dear Clubs, Life Members and Referees,<br><br>A quick reminder of this weekend\'s football across the '
               'association. This is a summary &mdash; tap any section heading or fixture to open the full, live detail on '
               f'<a href="{SITE}" target="_blank" style="color:{BLUE};font-weight:bold;text-decoration:none;">the website</a>, '
               'which is always the main source. Every weekend fixture is also attached as a spreadsheet for easy reference.')
        PREHEADER=(f'A quick reminder of this weekend\'s football — {sen} senior and {jun} junior fixtures '
                   f'({ncup} cup ties). Tap any section for the live detail on the website; the full fixtures list is attached.')
        FIX_TITLE="This weekend's fixtures"
        FIX_NOTE='Appointed match officials are shown on each card below where confirmed.'
        HEADLINE_NOTE='The email shows the Premier Division and cup ties.'
        FIX_ATTACH=(f'First Division, Second Division and Women\'s Premier, along with Reserves, 3rd–6th Divisions, Veterans and all '
                    f'<b style="color:{INK};">{jun} junior fixtures</b>, are on the website and in the <b style="color:{INK};">attached '
                    'spreadsheet</b> &mdash; open it in Excel and filter to your club to see all your games, home and away.')
        FIX_CTA='See all fixtures &amp; grounds on the website'
        RES_TITLE='Results & Standings'
        RES_HEAD='Latest results &middot; Premier'

    # ---- reusable blocks ----
    HEAD=f"""<!DOCTYPE html>
<html lang="en" xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><meta http-equiv="X-UA-Compatible" content="IE=edge">
<meta name="color-scheme" content="light only"><meta name="supported-color-schemes" content="light only"><title>CTTLFA {esc(KICKER)}</title>
<!--[if mso]><noscript><xml><o:OfficeDocumentSettings><o:PixelsPerInch>96</o:PixelsPerInch></o:OfficeDocumentSettings></xml></noscript><![endif]-->
<style>
 @media only screen and (max-width:1020px){{ .container{{width:100%!important}} }}
 @media only screen and (max-width:820px){{
   .px{{padding-left:24px!important;padding-right:24px!important}}
   .stack{{display:block!important;width:100%!important;padding-left:0!important;padding-right:0!important}} .stack+.stack{{margin-top:16px!important}}
   .col3{{display:block!important;width:100%!important;padding:0!important}}
   .sigl,.sigr{{display:block!important;width:100%!important}} .sigr{{padding-top:14px!important;border-left:0!important;padding-left:0!important}}
   .stat-n{{font-size:28px!important}} .h1{{font-size:27px!important}}
 }}
 a{{color:{BLUE}}} body{{margin:0;padding:0;background:#e9edf5}}
</style></head>
<body style="margin:0;padding:0;background:#e9edf5;-webkit-font-smoothing:antialiased;font-family:Arial,Helvetica,sans-serif;">
<div style="display:none;max-height:0;overflow:hidden;opacity:0;font-size:1px;line-height:1px;color:#e9edf5;">{PREHEADER}</div>
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#e9edf5;"><tr><td align="center" style="padding:26px 10px;">
<table role="presentation" class="container" width="1000" cellpadding="0" cellspacing="0" style="width:1000px;max-width:1000px;background:#ffffff;border-radius:18px;overflow:hidden;box-shadow:0 10px 34px rgba(7,26,74,.13);">

 <tr><td bgcolor="{NAVY}" style="background:{NAVY};padding:32px {PAD}px;" class="px"><table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
   <td width="62" valign="middle" style="width:62px;"><img src="{crest}" width="58" height="58" alt="Cape Town Tygerberg LFA crest" style="display:block;border:0;width:58px;height:58px;"></td>
   <td valign="middle" style="padding-left:18px;"><div style="color:#ffffff;font-size:21px;font-weight:bold;letter-spacing:.3px;line-height:1.15;">Cape Town Tygerberg LFA</div>
     <div style="color:{GOLD};font-size:12px;font-weight:bold;text-transform:uppercase;letter-spacing:2.2px;padding-top:5px;">{esc(KICKER)}</div></td>
   <td valign="middle" align="right" style="color:#9fb0d8;font-size:11px;font-weight:bold;text-transform:uppercase;letter-spacing:.8px;line-height:1.6;">Issue<br><span style="color:#ffffff;font-size:13px;">{ISSUE_LABEL}</span></td>
 </tr></table></td></tr>
 <tr><td bgcolor="{GOLD}" style="height:5px;background:{GOLD};font-size:0;line-height:0;">&nbsp;</td></tr>

 <tr><td class="px" style="padding:34px {PAD}px 0;">
   <div style="color:{MUT};font-size:12px;font-weight:bold;text-transform:uppercase;letter-spacing:1.5px;">{EYEBROW}</div>
   <div class="h1" style="color:{INK};font-size:32px;font-weight:bold;line-height:1.15;padding:10px 0 0;">{H1}</div>
   <p style="color:{MUT};font-size:15.5px;line-height:1.65;margin:14px 0 0;max-width:820px;">{INTRO}</p>
 </td></tr>

 <tr><td class="px" style="padding:18px {PAD}px 0;">
   <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#fbf7ec;border:1px solid #f0e2bf;border-radius:10px;"><tr>
     <td width="6" bgcolor="{GOLD}" style="width:6px;background:{GOLD};font-size:0;line-height:0;">&nbsp;</td>
     <td style="padding:13px 18px;font-size:12.5px;color:{INK};line-height:1.6;"><b style="color:{NAVY};">Snapshot notice.</b> This bulletin is a snapshot taken at <b>{STAMP}</b>. Fixtures, referee appointments and standings can change after it is sent. The website (<a href="{SITE}" target="_blank" style="color:{BLUE};font-weight:bold;text-decoration:none;">www.cttfa.co.za</a>) remains the single authoritative source &mdash; please confirm there before travelling.</td>
   </tr></table>
 </td></tr>

 <tr><td class="px" style="padding:24px {PAD}px 0;">
   <table role="presentation" width="100%" cellpadding="0" cellspacing="0" bgcolor="{NAVY}" style="background:{NAVY};border-radius:14px;"><tr>
     {stat(total,'Matches')}{stat(sen,'Senior')}{stat(jun,'Junior')}{stat(ncup,'Cup ties',True)}{stat(NCLUBS,'Clubs')}
   </tr></table>
 </td></tr>
"""

    FIXTURES=f"""
 <tr><td class="px" style="padding:36px {PAD}px 0;">{section_banner(FIX_TITLE, SITE+'/thisweekend')}
   <div style="color:{MUT};font-size:12.5px;padding-top:12px;"><span style="display:inline-block;background:{BLUE};color:#fff;font-size:10px;font-weight:bold;padding:2px 8px;border-radius:4px;">LEAGUE</span> &nbsp; <span style="display:inline-block;background:{GOLD};color:{INK};font-size:10px;font-weight:bold;padding:2px 8px;border-radius:4px;">&#9733; CUP / KNOCKOUT</span> &nbsp; {FIX_NOTE}</div>
 </td></tr>
 {days_html}
 <tr><td class="px" style="padding:14px {PAD}px 0;">
   <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:{SOFT};border-radius:10px;"><tr><td style="padding:16px 18px;color:{MUT};font-size:13.5px;line-height:1.65;">
     <b style="color:{INK};">{HEADLINE_NOTE}</b> {FIX_ATTACH}
   </td></tr></table></td></tr>
 <tr><td class="px" style="padding:16px {PAD}px 0;" align="center">
   <table role="presentation" cellpadding="0" cellspacing="0"><tr><td bgcolor="{BLUE}" style="border-radius:10px;"><a href="{SITE}/thisweekend" target="_blank" style="display:inline-block;font-size:15px;font-weight:bold;color:#ffffff;text-decoration:none;padding:13px 30px;border-radius:10px;">{FIX_CTA} &rarr;</a></td></tr></table>
 </td></tr>
"""

    if MODE=='weekahead':
        REFEREE=f"""
 <tr><td class="px" style="padding:36px {PAD}px 0;">{section_banner('Referee Appointments, Announcements & Rulings', 'https://dash.cttlfa.com/', cta='Open the dashboard')}
   <p style="color:{MUT};font-size:14.5px;line-height:1.65;margin:16px 0 0;">Referee appointments for this week&rsquo;s matches are confirmed midweek. They will appear on <b style="color:{INK};">every fixture in Thursday&rsquo;s weekend bulletin</b>, with the officials also carried in that bulletin&rsquo;s attached spreadsheet. The full appointment list for every division, along with association announcements and administrative rulings, is always live on the CTTLFA dashboard.</p></td></tr>
"""
    else:
        REFEREE=f"""
 <tr><td class="px" style="padding:36px {PAD}px 0;">{section_banner('Referee Appointments, Announcements & Rulings', 'https://dash.cttlfa.com/', cta='Open the dashboard')}
   <p style="color:{MUT};font-size:14.5px;line-height:1.65;margin:16px 0 0;">The appointed referee for each headline game is shown on the fixture cards above, and <b style="color:{INK};">every fixture in the attached spreadsheet carries its referee and assistants</b> &mdash; filter to your club to see the officials for all your games. The full appointment list for every division, along with association announcements and administrative rulings, is on the CTTLFA dashboard. Appointments marked <b style="color:{INK};">TBA</b> are still to be confirmed, and officials can change &mdash; the dashboard is the live source.</p></td></tr>
"""

    RESULTS=f"""
 <tr><td class="px" style="padding:40px {PAD}px 0;">{section_banner(RES_TITLE, SITE+'/matchcentre')}</td></tr>
 <tr><td class="px" style="padding:18px {PAD}px 0;"><table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
   <td class="stack" width="50%" valign="top" style="padding-right:18px;">
     <div style="color:{NAVY};font-size:14px;font-weight:bold;padding-bottom:8px;">{RES_HEAD}</div>
     <table role="presentation" width="100%" cellpadding="0" cellspacing="0">{results_html}</table></td>
   <td class="stack" width="50%" valign="top" style="padding-left:18px;">
     <div style="color:{NAVY};font-size:14px;font-weight:bold;padding-bottom:8px;">Standings &middot; Premier</div>
     <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:14px;color:{INK};">
       <tr style="color:{MUT};font-size:11px;text-transform:uppercase;"><td style="padding:0 0 4px;">#</td><td style="padding:0 0 4px;">Club</td><td align="center" style="padding:0 0 4px;">GD</td><td align="right" style="padding:0 0 4px;">Pts</td></tr>{ptable_html}
     </table>
     <div style="color:{MUT};font-size:12.5px;padding-top:10px;">First Division leaders: <b style="color:{INK};">{esc(first_lead)}</b> &nbsp;&middot;&nbsp; Second Division leaders: <b style="color:{INK};">{esc(second_lead)}</b></div></td>
 </tr></table></td></tr>
"""

    JLOGS=f"""
 <tr><td class="px" style="padding:40px {PAD}px 0;">{section_banner('Junior Combined Logs', SITE+'/combinedlogs')}
   <div style="color:{MUT};font-size:13px;padding-top:12px;">Each club's Under-12 to Under-18 teams added together and ranked as one. <span style="color:{GNT};font-weight:bold;">&#9632; Promotion</span> &nbsp; <span style="color:{RDT};font-weight:bold;">&#9632; Relegation</span></div></td></tr>
 <tr><td class="px" style="padding:18px {PAD}px 0;"><table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
   <td class="stack" width="50%" valign="top" style="padding-right:14px;">{logs_table(*TIERS[0])}</td>
   <td class="stack" width="50%" valign="top" style="padding-left:14px;">{logs_table(*TIERS[1])}</td></tr></table></td></tr>
 <tr><td class="px" style="padding:24px {PAD}px 0;"><table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
   <td class="stack" width="50%" valign="top" style="padding-right:14px;">{logs_table(*TIERS[2])}</td>
   <td class="stack" width="50%" valign="top" style="padding-left:14px;">{logs_table(*TIERS[3])}</td></tr></table></td></tr>
 <tr><td class="px" style="padding:18px {PAD}px 0;"><table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#fbf7ec;border:1px solid #f0e2bf;border-radius:10px;"><tr><td style="padding:13px 16px;font-size:12px;color:{MUT};line-height:1.6;font-style:italic;">Disclaimer: standings are subject to change pending the outcomes of any outstanding Disciplinary Committee matters and decisions. Full Under-12 to Under-18 age-group breakdowns are on the website.</td></tr></table></td></tr>
"""

    SIGNOFF=f"""
 <tr><td class="px" style="padding:40px {PAD}px 0;">
   <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border:1px solid {LINE};border-left:5px solid {GOLD};border-radius:12px;background:#ffffff;"><tr><td style="padding:24px 28px;">
     <div style="color:{MUT};font-size:14px;">Kind regards,</div>
     <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-top:6px;"><tr>
       <td class="sigl" valign="top" width="46%"><div style="color:{NAVY};font-size:19px;font-weight:bold;">Barry Petersen</div>
         <div style="color:{MUT};font-size:13px;padding-top:3px;">Operations Manager</div><div style="color:{MUT};font-size:13px;">Cape Town Tygerberg LFA</div></td>
       <td class="sigr" valign="top" width="54%" style="border-left:1px solid {LINE};padding-left:26px;">
         <table role="presentation" cellpadding="0" cellspacing="0" style="font-size:13px;line-height:2;">
           <tr><td style="color:{MUT};text-transform:uppercase;font-size:10.5px;letter-spacing:.6px;padding-right:16px;">Office</td><td style="color:{INK};">+27 21 686 4004</td></tr>
           <tr><td style="color:{MUT};text-transform:uppercase;font-size:10.5px;letter-spacing:.6px;padding-right:16px;">Mobile</td><td style="color:{INK};">+27 84 964 3951</td></tr>
           <tr><td style="color:{MUT};text-transform:uppercase;font-size:10.5px;letter-spacing:.6px;padding-right:16px;">Email</td><td><a href="mailto:ops@cttlfa.com" style="color:{BLUE};text-decoration:none;">ops@cttlfa.com</a></td></tr>
           <tr><td style="color:{MUT};text-transform:uppercase;font-size:10.5px;letter-spacing:.6px;padding-right:16px;">Web</td><td><a href="{SITE}" style="color:{BLUE};text-decoration:none;">www.cttfa.co.za</a></td></tr>
         </table></td>
     </tr></table></td></tr></table></td></tr>

 <tr><td class="px" style="padding:26px {PAD}px 30px;" align="center"><div style="font-size:13px;color:{BLUE};line-height:2.2;">
   <a href="{SITE}/thisweekend" style="color:{BLUE};text-decoration:none;font-weight:bold;">Fixtures</a> &nbsp;&middot;&nbsp; <a href="{SITE}/matchcentre" style="color:{BLUE};text-decoration:none;font-weight:bold;">Results &amp; Tables</a> &nbsp;&middot;&nbsp; <a href="{SITE}/knockouts" style="color:{BLUE};text-decoration:none;font-weight:bold;">Knockout Cups</a> &nbsp;&middot;&nbsp; <a href="{SITE}/combinedlogs" style="color:{BLUE};text-decoration:none;font-weight:bold;">Junior Logs</a> &nbsp;&middot;&nbsp; <a href="{SITE}/grounds" style="color:{BLUE};text-decoration:none;font-weight:bold;">Grounds</a></div></td></tr>

 <tr><td bgcolor="{NAVY}" style="background:{NAVY};padding:28px {PAD}px;" class="px" align="center">
   <div style="color:{GOLD};font-size:13px;font-weight:bold;letter-spacing:.4px;">Cape Town Tygerberg Local Football Association</div>
   <div style="color:#9fb0d8;font-size:11.5px;line-height:1.8;padding-top:8px;">Established 1993 &middot; Cape Town, South Africa<br>This weekly bulletin is generated automatically from the live association website.<br>You are receiving it as an affiliated club, life member or match official of the CTTLFA.</div>
   <div style="padding-top:10px;"><a href="mailto:ops@cttlfa.com" style="color:#fff;font-size:11.5px;text-decoration:underline;">Contact the office</a></div>
 </td></tr>

</table></td></tr></table></body></html>"""

    if MODE=='weekahead':
        body=RESULTS+FIXTURES+REFEREE+JLOGS
    else:
        body=FIXTURES+REFEREE+RESULTS+JLOGS
    return HEAD+body+SIGNOFF

def _textver():
    if MODE=='weekahead':
        L=['CAPE TOWN TYGERBERG LFA - THE WEEK AHEAD','Week: %s'%RANGE_LABEL,'',
           'Snapshot at %s. The website (%s) remains the authoritative source.'%(STAMP,SITE),'',
           'This week: %d matches (%d senior, %d junior, %d cup ties) across %d clubs, Monday to Sunday.'%(total,sen,jun,ncup,NCLUBS),'',
           'Every fixture this week is attached as a spreadsheet - filter to your club.',
           'Referee appointments are confirmed midweek and will appear in Thursday\'s weekend bulletin.','',
           'Fixtures: %s/thisweekend'%SITE,'Results & Tables: %s/matchcentre'%SITE,'Junior Logs: %s/combinedlogs'%SITE,'',
           'Kind regards,','Barry Petersen - Operations Manager, Cape Town Tygerberg LFA','ops@cttlfa.com']
    else:
        L=['CAPE TOWN TYGERBERG LFA - WEEKLY CLUB BULLETIN','Match Week: %s'%RANGE_LABEL,'',
           'Snapshot at %s. The website (%s) remains the authoritative source.'%(STAMP,SITE),'',
           'This weekend: %d matches (%d senior, %d junior, %d cup ties) across %d clubs.'%(total,sen,jun,ncup,NCLUBS),'',
           'Every weekend fixture (with referee & assistants where confirmed) is attached as a spreadsheet - filter to your club.',
           'Fixtures: %s/thisweekend'%SITE,'Results & Tables: %s/matchcentre'%SITE,'Junior Logs: %s/combinedlogs'%SITE,'',
           'Kind regards,','Barry Petersen - Operations Manager, Cape Town Tygerberg LFA','ops@cttlfa.com']
    return chr(10).join(L)

def _send(html_out, xlsx_path):
    KEY=os.environ.get('RESEND_API_KEY','').strip()
    TO=[x.strip() for x in re.split(r'[,;\s]+', os.environ.get('RESEND_TO','')) if x.strip()]
    FROM=os.environ.get('RESEND_FROM') or 'Cape Town Tygerberg LFA <bulletin@cttlfa.com>'
    if not KEY or not TO:
        print('DRY RUN: RESEND_API_KEY/RESEND_TO not set - files built, no email sent.'); return 0
    with open(xlsx_path,'rb') as f: att=base64.b64encode(f.read()).decode()
    subject=('CTTLFA The Week Ahead - %s' if MODE=='weekahead' else 'CTTLFA Weekly Club Bulletin - %s')%RANGE_LABEL
    fname=('CTTLFA Week Ahead Fixtures %s.xlsx' if MODE=='weekahead' else 'CTTLFA Weekend Fixtures %s.xlsx')%RANGE_LABEL
    sent=0
    for rcpt in TO:  # one email per club: no address leakage between recipients
        payload={'from':FROM,'to':[rcpt],'reply_to':'ops@cttlfa.com','subject':subject,'html':html_out,'text':_textver(),
                 'headers':{'List-Unsubscribe':'<mailto:ops@cttlfa.com?subject=Unsubscribe%20CTTLFA%20bulletin>'},
                 'attachments':[{'filename':fname,'content':att}]}
        req=urllib.request.Request('https://api.resend.com/emails',
            data=json.dumps(payload).encode(),
            headers={'Authorization':'Bearer '+KEY,'Content-Type':'application/json',
                     'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'})
        try:
            r=urllib.request.urlopen(req,timeout=60).read().decode(); sent+=1
            print('sent ->',rcpt,r[:80])
        except Exception as e:
            body=''
            try: body=e.read().decode('utf-8','ignore')[:400]
            except Exception: pass
            print('FAILED ->',rcpt,e,'| FROM=',FROM,'| BODY=',body)
    return sent

def main():
    print('Mode %s | run %s | range %s | season %s | mapping %s'%(MODE,RUN,RANGE_LABEL,SEASON_SRC,MAP_SRC))
    if total==0:
        print('No fixtures for the %s - silent skip (off-season/bye).'%('week' if MODE=='weekahead' else 'coming weekend')); return
    outdir=os.environ.get('OUT_DIR','downloads'); os.makedirs(outdir,exist_ok=True)
    stem=('week-ahead-fixtures' if MODE=='weekahead' else 'weekend-fixtures')
    htmlname=('week-ahead-bulletin.html' if MODE=='weekahead' else 'weekly-bulletin.html')
    datestamp=(_WD[0] if MODE=='weekahead' else _SAT).isoformat()
    xlsx_path=os.path.join(outdir,stem+'.xlsx')
    nx=make_xlsx(xlsx_path); print('xlsx: %d club-rows / %d matches'%nx)
    shutil.copyfile(xlsx_path, os.path.join(outdir,'%s-%s.xlsx'%(stem,datestamp)))
    html_out=page(SITE+'/assets/crest.png')
    open(os.path.join(outdir,htmlname),'w',encoding='utf-8').write(html_out)
    print('built HTML (%d bytes) + Excel in %s/'%(len(html_out),outdir))
    n=_send(html_out, xlsx_path)
    print('emails sent: %d'%n)

if __name__=='__main__':
    main()
